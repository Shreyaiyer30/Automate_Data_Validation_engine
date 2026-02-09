import streamlit as st
import pandas as pd
import json
import io

def render_export_section(df: pd.DataFrame, report: dict):
    """Renders export options for data and reports."""
    st.markdown("## 📥 EXPORT & FINAL REPORT")
    
    # 1. Summary Score Card
    c_score, c_ret = st.columns(2)
    with c_score:
        st.markdown(f"""
        <div class='glass-card' style='text-align: center; border-top: 4px solid #00ffff;'>
            <h1 style='color: #ffffff;'>{report['quality_score']}%</h1>
            <p style='color: #888;'>Final Data Quality Score</p>
        </div>
        """, unsafe_allow_html=True)
    with c_ret:
        st.markdown(f"""
        <div class='glass-card' style='text-align: center; border-top: 4px solid #ff00ff;'>
            <h1 style='color: #ffffff;'>{report['summary']['retention_rate']}%</h1>
            <p style='color: #888;'>Row Retention Rate</p>
        </div>
        """, unsafe_allow_html=True)

    # 2. Detailed Impact Summary
    st.markdown("### 📊 Impact Summary")
    s = report.get('summary', {})
    stats = report.get('statistics', {})
    
    col_sum1, col_sum2, col_sum3 = st.columns(3)
    col_sum1.metric("Rows Preserved", f"{stats.get('final', {}).get('rows', 0):,}")
    col_sum2.metric("Data Mutations", f"{s.get('total_actions', 0):,}")
    col_sum3.metric("Blocked Rows", f"{s.get('rows_removed', 0):,}")

    # 3. Validation Results (Enterprise Structure)
    st.markdown("### ✅ Validation Status")
    
    # We derive categorization from report issues and audit log
    warnings = [e for e in report.get('audit_log', []) if e.get('state') == 'WARN']
    failures = [e for e in report.get('audit_log', []) if e.get('state') == 'FAIL']
    
    v_pass, v_warn, v_fail = st.tabs(["🟢 Passed Checks", "🟡 Warnings", "🔴 Failures"])
    
    with v_pass:
        st.success("All structural and schema integrity checks passed.")
    
    with v_warn:
        if warnings:
            for w in warnings:
                st.warning(f"**{w.get('stage')}**: {w.get('message', 'Issue detected during execution.')}")
        else:
            st.info("No warnings issued during the pipeline run.")
            
    with v_fail:
        if failures:
            for f in failures:
                st.error(f"**{f.get('stage')}**: {f.get('message', 'Critical failure.')}")
        else:
            st.info("No blocking failures detected. Pipeline integrity preserved.")

    # 4. Downloads (Dynamic Presentation Layer)
    st.markdown("### 📥 Download Results")
    c1, c2, c3 = st.columns(3)
    

    # Extract original basename for naming contract
    metadata = st.session_state.get('metadata', {})
    raw_filename = metadata.get('original_filename', 'data') if isinstance(metadata, dict) else getattr(metadata, 'original_filename', 'data')
    base_name = raw_filename.rsplit('.', 1)[0] if '.' in raw_filename else raw_filename
    
    with c1:
        # CSV Export (Headers already normalized in Phase 4)
        if hasattr(st.session_state.processor, 'audit_export'):
            st.session_state.processor.audit_export("CSV", list(df.columns))
        st.download_button("📥 DOWNLOAD CSV", 
                          df.to_csv(index=False).encode('utf-8'),
                          f"{base_name}_cleaned_data.csv", "text/csv", use_container_width=True)
                          
    with c2:
        # Rich Excel Export
        output = io.BytesIO()
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            if hasattr(st.session_state.processor, 'audit_export'):
                st.session_state.processor.audit_export("EXCEL", list(df.columns))
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Cleaned Data')
                
                # Access sheet for rich formatting
                worksheet = writer.sheets['Cleaned Data']
                row_count, col_count = df.shape
                
                # 1. Convert to Excel Table (Filters + Striping)
                # Note: displayName must be alphanumeric
                tab_name = "CleanedData_" + base_name.replace(" ", "_")[:20]
                ref = f"A1:{get_column_letter(col_count)}{row_count + 1}"
                tab = Table(displayName=tab_name, ref=ref)
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                tab.tableStyleInfo = style
                worksheet.add_table(tab)
                
                # 2. Freeze Panes (Header)
                worksheet.freeze_panes = 'A2'
                
                # 3. Dynamic Column Widths (Auto-fit)
                for i, column in enumerate(df.columns):
                    column_len = max(df[column].astype(str).map(len).max(), len(column))
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = min(column_len + 2, 50)
                    
        except ImportError:
            # Fallback if openpyxl features fail / aren't fully available
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                
        excel_data = output.getvalue()
        st.download_button("📥 DOWNLOAD EXCEL",
                          excel_data,
                          f"{base_name}_cleaned_data.xlsx", 
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                          use_container_width=True)
                          
    with c3:
        st.download_button("📥 DOWNLOAD JSON REPORT",
                          json.dumps(report, indent=2).encode('utf-8'),
                          f"{base_name}_quality_report.json", "application/json", use_container_width=True)
    
    # Audit log preview
    with st.expander("📝 View Pipeline Audit History", expanded=True):
        audit_data = report.get('audit_log', [])
        if audit_data:
            audit_df = pd.DataFrame(audit_data)
            
            # Reorder for clarity if columns exist
            cols = [c for c in ['timestamp', 'stage', 'state', 'message', 'details'] if c in audit_df.columns]
            audit_df = audit_df[cols]
            
            # Styling Logic (Severity-based)
            def color_severity(val):
                if str(val).upper() == 'FAIL': return 'background-color: rgba(255, 0, 0, 0.2); font-weight: bold;'
                if str(val).upper() == 'WARN': return 'background-color: rgba(255, 165, 0, 0.2);'
                if str(val).upper() in ['PASS', 'SUCCESS']: return 'background-color: rgba(0, 255, 0, 0.1);'
                return ''
            
            styled_audit = audit_df.style.applymap(color_severity, subset=['state'] if 'state' in audit_df.columns else [])
            
            # Dynamic Headers
            audit_cfg = {col: col.replace('_', ' ').title() for col in audit_df.columns}
            
            st.dataframe(styled_audit, use_container_width=True, column_config=audit_cfg, hide_index=True)
        else:
            st.info("No audit events were captured for this session.")
        
        st.markdown("---")
        st.caption("Full JSON Metadata")
        st.json(report)
